import streamlit as st
import pandas as pd
import os
from sqlmodel import Field, Session, SQLModel, create_engine, select
from sqlalchemy.exc import InvalidRequestError
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from io import StringIO
from typing import get_args, get_origin, Union, Optional
import json
from pydantic import create_model
import importlib.util
import keyword
import re


def get_db_engine(db_name):
    # Create a database engine
    db_dir = "databases"
    if not os.path.exists(db_dir):
        os.makedirs(db_dir)

    db_file = os.path.join(db_dir, f"{db_name}.db")
    if "engines" not in st.session_state:
        st.session_state["engines"] = {}
    if db_name not in st.session_state["engines"]:
        st.session_state["engines"][db_name] = create_engine(f"sqlite:///{db_file}")
    return st.session_state["engines"][db_name]


def import_generated_models():
    """
    Imports the generated model classes from the models directory.
    """
    models_dir = "models"
    if not st.session_state.get("imported_db_classes") or not os.path.exists(
        models_dir
    ):
        st.session_state["imported_db_classes"] = {}
    else:
        if "engines" not in st.session_state:
            st.session_state["engines"] = {}
        model_files = [f for f in os.listdir(models_dir) if f.endswith(".py")]
        for model_file_name in model_files:
            if model_file_name.endswith("__init__.py"):
                continue
            db_name = model_file_name.replace(".py", "")
            model_name = f"{db_name}_model"
            db_path = os.path.join(models_dir, model_file_name)
            if model_name not in st.session_state["imported_db_classes"]:
                st.session_state["imported_db_classes"][model_name] = (
                    import_model_class(db_path, model_name)
                )
            if model_name not in st.session_state["engines"]:
                st.session_state["engines"][model_name] = get_db_engine(db_name)


if "imported_db_classes" not in st.session_state:
    import_generated_models()

## ADD Engine to the Session State to avoid multiple connections


def generate_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Schema"

    # Add instructions
    ws["A1"] = (
        "Instructions: Fill out the field names and select data types from the list."
    )
    ws.merge_cells("A1:B1")
    ws["A2"] = "Field Name"
    ws["B2"] = "Data Type"

    # Provide some placeholder entries
    ws["A3"] = "id"
    ws["B3"] = "integer"
    ws["A4"] = "name"
    ws["B4"] = "string"

    # Define the data type list
    data_types = ["string", "integer", "float", "date"]
    data_type_str = ",".join(data_types)

    # Create a data validation object
    dv = DataValidation(type="list", formula1=f'"{data_type_str}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("B3:B1048576")  # Apply to column B

    # Protect the header row to prevent edits

    # Save the workbook to a BytesIO stream
    wb.save("database_schema_template.xlsx")

    return "database_schema_template.xlsx"


def create_pydantic_model(schema_df, model_name):
    # Prepare fields for the Pydantic model
    fields = {}
    for index, row in schema_df.iterrows():
        field_name = row["Field Name"].strip()
        data_type = row["Data Type"].lower().strip()

        if data_type == "string":
            field_type = (Optional[str], None)
        elif data_type == "integer":
            field_type = (Optional[int], None)
        elif data_type == "float":
            field_type = (Optional[float], None)
        elif data_type == "date":
            field_type = (
                Optional[str],
                None,
            )  # Dates can be handled as strings or date objects
        else:
            st.error(f"Unsupported data type: {data_type}")
            continue

        fields[field_name] = field_type

    # Create Pydantic model
    pydantic_model = create_model(model_name, **fields)
    return pydantic_model


def save_model_to_file(pydantic_model, model_name, db_name, primary_keys):
    """
    Saves the Pydantic model as a SQLModel class to a Python file.
    Ensures that field names are valid Python identifiers and that
    type hints are correctly mapped.

    Args:
        pydantic_model: The Pydantic model to be converted.
        model_name: Name of the model class.
        db_name: Name of the database.
        primary_keys: List of primary key fields.

    Returns:
        The path to the model file.
    """
    # Directory to save model files
    models_dir = "models"
    if not os.path.exists(models_dir):
        os.makedirs(models_dir)

    model_file = os.path.join(models_dir, f"{db_name}.py")

    # Start building the model code
    model_code = "from sqlmodel import Field, SQLModel\n"
    model_code += "from typing import Optional\n\n"
    model_code += f"class {model_name}(SQLModel, table=True):\n"

    # Process each field in the Pydantic model
    for field_name, field_type in pydantic_model.__annotations__.items():
        # Sanitize the field name
        sanitized_field_name = sanitize_field_name(field_name)
        if sanitized_field_name != field_name:
            st.warning(
                f"Field name '{field_name}' has been changed to '{sanitized_field_name}' to be a valid Python identifier."
            )
            field_name = sanitized_field_name

        # Get the type hint as a string
        type_hint = get_type_hint(field_type)

        # Check if the field is a primary key
        is_primary_key = field_name in primary_keys

        # Build the field definition
        if is_primary_key:
            field_def = f"    {field_name}: {type_hint} = Field(primary_key=True)"
        else:
            field_def = f"    {field_name}: {type_hint} = Field(default=None)"

        # Add the field definition to the model code
        model_code += f"{field_def}\n"

    # Save the model code to the file
    with open(model_file, "w") as f:
        f.write(model_code)

    return model_file


def sanitize_field_name(field_name):
    """
    Sanitizes the field name to ensure it is a valid Python identifier.

    Args:
        field_name: The original field name.

    Returns:
        A sanitized field name that is a valid Python identifier.
    """
    # Remove leading/trailing whitespace and replace spaces with underscores
    field_name = field_name.strip().replace(" ", "_")

    # Replace invalid characters with underscores
    field_name = re.sub(r"\W|^(?=\d)", "_", field_name)

    # Check if the field name is a Python keyword
    if keyword.iskeyword(field_name):
        field_name += "_field"

    # Ensure the field name is not empty
    if not field_name:
        field_name = "field"

    return field_name


def normalize_db_name(db_name: str):
    return (
        db_name.strip()
        .replace("_", " ")
        .title()
        .replace("_db", "")
        .replace("_model", "")
        .strip()
    )


def get_type_hint(field_type):
    """
    Converts a field type to a string suitable for type hints in code.

    Args:
        field_type: The original field type from the Pydantic model.

    Returns:
        A string representing the type hint.
    """
    origin = get_origin(field_type)
    args = get_args(field_type)

    if origin is Union and type(None) in args:
        # It's an Optional type
        actual_type = [arg for arg in args if arg is not type(None)][0]
        actual_type_str = get_type_str(actual_type)
        return f"Optional[{actual_type_str}]"
    else:
        return get_type_str(field_type)


def get_type_str(field_type):
    """
    Maps a Python type to its string representation for code generation.

    Args:
        field_type: The field type.

    Returns:
        A string representing the type.
    """
    if field_type is int:
        return "int"
    elif field_type is float:
        return "float"
    elif field_type is str:
        return "str"
    elif field_type is bool:
        return "bool"
    else:
        # Default to 'str' for unsupported types
        return "str"


def import_model_class(model_file_path, model_name):
    # Import the model class from the model file
    spec = importlib.util.spec_from_file_location(model_name, model_file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    model_class = getattr(module, model_name)

    # Handle table already defined error
    try:
        SQLModel.metadata.create_all(
            get_db_engine(db_name=model_name.replace("_model", ""))
        )  # Create tables if not already created
    except InvalidRequestError as e:
        if "already defined" in str(e):
            print(
                f"Table '{model_name}' is already defined. Returning the existing model class."
            )
        else:
            raise e  # If it's another error, re-raise the exception

    return model_class


def generate_database(db_name, schema_df, primary_keys, description):
    # Create Pydantic model
    model_name = f"{db_name}_model"
    pydantic_model = create_pydantic_model(schema_df, model_name)

    # Save model to file and convert to SQLModel
    model_file = save_model_to_file(pydantic_model, model_name, db_name, primary_keys)

    # Import the SQLModel class from the file
    if model_name in st.session_state["imported_db_classes"]:
        print(model_name, "found in session state.")
        model_class = st.session_state["imported_db_classes"][model_name]
    else:
        print(model_name, "not found in session state. Importing from file.")
        model_class = import_model_class(model_file, model_name)
        st.session_state["imported_db_classes"][model_name] = model_class

    # Create the database
    # Save the schema to a JSON file
    schemas_dir = "schemas"
    if not os.path.exists(schemas_dir):
        os.makedirs(schemas_dir)

    schema_json = schema_df.to_json()
    schema_file = os.path.join(schemas_dir, f"{db_name}_schema.json")
    with open(schema_file, "w") as f:
        f.write(schema_json)

    # Save primary keys
    pk_file = os.path.join(schemas_dir, f"{db_name}_pk.json")
    with open(pk_file, "w") as f:
        json.dump(primary_keys, f)

    # Save description
    desc_file = os.path.join(schemas_dir, f"{db_name}_desc.txt")
    with open(desc_file, "w") as f:
        f.write(description)

    return model_class


def interact_with_database(db_name):
    st.header(f"Interact with Database: {normalize_db_name(db_name)}")

    # Load the schema
    schemas_dir = "schemas"
    schema_file = os.path.join(schemas_dir, f"{db_name}_schema.json")
    pk_file = os.path.join(schemas_dir, f"{db_name}_pk.json")
    desc_file = os.path.join(schemas_dir, f"{db_name}_desc.txt")

    if not os.path.exists(schema_file):
        st.error("Schema file not found.")
        return
    if not os.path.exists(pk_file):
        st.error("Primary key file not found.")
        return

    with open(schema_file, "r") as f:
        schema_json = f.read()
    schema_df = pd.read_json(StringIO(schema_json))

    with open(pk_file, "r") as f:
        primary_keys = json.load(f)

    if os.path.exists(desc_file):
        with open(desc_file, "r") as f:
            description = f.read()
    else:
        description = "No description provided."

    # Import the model class from the model file
    models_dir = "models"
    model_file = os.path.join(models_dir, f"{db_name}.py")
    model_name = f"{db_name}_model"
    if model_name in st.session_state["imported_db_classes"]:
        print(model_name, "found in session state.")
        model_class = st.session_state["imported_db_classes"][model_name]
        engine = get_db_engine(db_name)
    else:
        print(model_name, "not found in session state. Importing from file.")
        model_class = import_model_class(model_file, model_name)
        engine = get_db_engine(db_name)
    # Create engine

    # Database Details
    st.subheader("Database Details")
    st.write(f"**Name**: {normalize_db_name(db_name)}")
    st.write(f"**Description**: {description}")

    # Number of records, etc.
    with Session(engine) as session:
        statement = select(model_class)
        results = session.exec(statement)
        records = results.all()
        record_count = len(records)
        st.write(f"**Number of records**: {record_count}")

    # CRUD Forms
    st.subheader("CRUD Operations")

    # Create
    st.write("### Create New Record")
    with st.form("create_form"):
        form_fields = {}
        for field_name, field_type in model_class.__annotations__.items():
            if field_type == Optional[int] or field_type == int:
                form_fields[field_name] = st.number_input(field_name, step=1)
            elif field_type == Optional[float] or field_type == float:
                form_fields[field_name] = st.number_input(field_name)
            else:
                form_fields[field_name] = st.text_input(field_name)
        if st.form_submit_button("Add Record"):
            try:
                new_record_data = {}
                for field_name, field_value in form_fields.items():
                    field_type = model_class.__annotations__[field_name]
                    if field_type == Optional[int] or field_type == int:
                        new_record_data[field_name] = int(field_value)
                    elif field_type == Optional[float] or field_type == float:
                        new_record_data[field_name] = float(field_value)
                    else:
                        new_record_data[field_name] = field_value
                new_record = model_class(**new_record_data)
                with Session(engine) as session:
                    session.add(new_record)
                    session.commit()
                    st.success("Record added successfully!")
                    st.rerun()
            except Exception as e:
                st.error(f"Error adding record: {e}")

    # Read
    st.write("### View Records")
    with Session(engine) as session:
        statement = select(model_class)
        results = session.exec(statement)
        records = results.all()
        df = pd.DataFrame([record.dict() for record in records])
        st.dataframe(df)

    # Update
    st.write("### Update Record")
    if not records:
        st.write("No records to update.")
    else:
        primary_key_field = primary_keys[0]
        record_ids = [getattr(record, primary_key_field) for record in records]
        selected_id = st.selectbox("Select Record ID to Update:", options=record_ids)
        if selected_id:
            with Session(engine) as session:
                record = session.get(model_class, selected_id)
                with st.form("update_form"):
                    update_fields = {}
                    for field_name, field_value in record.dict().items():
                        if field_name in primary_keys:
                            update_fields[field_name] = st.text_input(
                                field_name, value=str(field_value), disabled=True
                            )
                        else:
                            field_type = model_class.__annotations__[field_name]
                            if field_type == Optional[int] or field_type == int:
                                update_fields[field_name] = st.number_input(
                                    field_name, value=int(field_value), step=1
                                )
                            elif field_type == Optional[float] or field_type == float:
                                update_fields[field_name] = st.number_input(
                                    field_name, value=float(field_value)
                                )
                            else:
                                update_fields[field_name] = st.text_input(
                                    field_name, value=str(field_value)
                                )
                    if st.form_submit_button("Update Record"):
                        try:
                            for field_name, field_value in update_fields.items():
                                field_type = model_class.__annotations__[field_name]
                                if field_type == Optional[int] or field_type == int:
                                    setattr(record, field_name, int(field_value))
                                elif (
                                    field_type == Optional[float] or field_type == float
                                ):
                                    setattr(record, field_name, float(field_value))
                                else:
                                    setattr(record, field_name, field_value)
                            session.add(record)
                            session.commit()
                            st.success("Record updated successfully!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error updating record: {e}")

    # Delete
    st.write("### Delete Record")
    if not records:
        st.write("No records to delete.")
    else:
        delete_id = st.selectbox(
            "Select Record ID to Delete:", options=record_ids, key="delete_select"
        )
        if st.button("Delete Record"):
            try:
                with Session(engine) as session:
                    record = session.get(model_class, delete_id)
                    if record:
                        session.delete(record)
                        session.commit()
                        st.success("Record deleted successfully!")
                        st.rerun()
            except Exception as e:
                st.error(f"Error deleting record: {e}")


def main():
    st.title("Dynamic Database Generator")

    # Excel Template Creation and Handling
    st.header("Step 1: Download Excel Template")
    st.write("Generate and download the Excel template to define your database schema.")

    if st.button("Download Excel Template"):
        template_path = generate_excel_template()
        with open(template_path, "rb") as f:
            st.download_button(
                label="Download Template",
                data=f,
                file_name="database_schema_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # Allow users to upload the completed template
    st.header("Step 2: Upload Completed Template")
    uploaded_file = st.file_uploader(
        "Upload the completed Excel template:", type=["xlsx"]
    )

    if uploaded_file is not None:
        # Parse the uploaded Excel file
        schema_df = pd.read_excel(uploaded_file, header=1)

        # Display the extracted schema
        st.header("Extracted Schema")
        st.write("Verify and customize your schema below.")
        st.dataframe(schema_df)

        # Allow users to modify field names and data types
        st.write("Modify field names and select data types:")
        edited_df = st.data_editor(schema_df)
        edited_df.columns = edited_df.columns.str.strip()
        st.write("Column Names:", edited_df.columns.tolist())

        # Proceed to database attributes input
        st.header("Step 3: Input Database Attributes")
        try:
            field_names = edited_df["Field Name"].tolist()
        except KeyError:
            st.error("The uploaded file does not contain a 'Field Name' column.")
            st.stop()
        db_name = st.text_input("Database Name", value="my_database")
        db_name = db_name.strip() + "_db"
        db_name = sanitize_field_name(db_name)
        description = st.text_area(
            "Description", value="A brief description of the database."
        )

        # Allow users to select primary keys
        primary_keys = st.multiselect(
            "Select Primary Key(s)",
            options=field_names,
            default=field_names[0] if field_names else None,
        )

        if st.button("Generate Database"):
            with st.spinner("Generating database..."):
                try:
                    generate_database(db_name, edited_df, primary_keys, description)
                    st.success("Database generated successfully!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error generating database: {e}")

    # Database Selection Interface
    st.header("Step 4: Select a Database to Interact With")
    db_dir = "databases"
    if os.path.exists(db_dir):
        db_files = [f for f in os.listdir(db_dir) if f.endswith(".db")]
        db_names = [os.path.splitext(f)[0] for f in db_files]
        if db_names:
            selected_db = st.selectbox("Select a database:", options=db_names)

            if selected_db:
                # Provide options to delete or rename databases
                st.write(
                    f"Selected Database: {normalize_db_name(selected_db)} ({selected_db})"
                )
                if st.button("Delete Database"):
                    confirm_delete = st.checkbox(
                        "Are you sure you want to delete this database?"
                    )
                    if confirm_delete:
                        os.remove(os.path.join(db_dir, f"{selected_db}.db"))
                        # Also remove schema and model files
                        schemas_dir = "schemas"
                        models_dir = "models"
                        schema_file = os.path.join(
                            schemas_dir, f"{selected_db}_schema.json"
                        )
                        pk_file = os.path.join(schemas_dir, f"{selected_db}_pk.json")
                        desc_file = os.path.join(schemas_dir, f"{selected_db}_desc.txt")
                        model_file = os.path.join(models_dir, f"{selected_db}.py")
                        if os.path.exists(schema_file):
                            os.remove(schema_file)
                        if os.path.exists(pk_file):
                            os.remove(pk_file)
                        if os.path.exists(desc_file):
                            os.remove(desc_file)
                        if os.path.exists(model_file):
                            os.remove(model_file)
                        st.success(f"Database {selected_db} deleted.")
                        st.rerun()
                rename_db = st.text_input(
                    "Enter new database name:", value=normalize_db_name(selected_db)
                )
                if st.button("Rename Database"):
                    if (
                        rename_db
                        and rename_db != selected_db
                        and rename_db != sanitize_field_name(selected_db) + "_db"
                    ):
                        clean_rename_db = sanitize_field_name(rename_db)
                        os.rename(
                            os.path.join(db_dir, f"{selected_db}.db"),
                            os.path.join(
                                db_dir, f"{sanitize_field_name(clean_rename_db)}.db"
                            ),
                        )
                        # Also rename schema and model files
                        schemas_dir = "schemas"
                        models_dir = "models"
                        os.rename(
                            os.path.join(schemas_dir, f"{selected_db}_schema.json"),
                            os.path.join(schemas_dir, f"{clean_rename_db}_schema.json"),
                        )
                        os.rename(
                            os.path.join(schemas_dir, f"{selected_db}_pk.json"),
                            os.path.join(schemas_dir, f"{clean_rename_db}_pk.json"),
                        )
                        os.rename(
                            os.path.join(schemas_dir, f"{selected_db}_desc.txt"),
                            os.path.join(schemas_dir, f"{clean_rename_db}_desc.txt"),
                        )
                        os.rename(
                            os.path.join(models_dir, f"{selected_db}_model.py"),
                            os.path.join(models_dir, f"{clean_rename_db}_model.py"),
                        )
                        st.success(
                            f"Database {selected_db} renamed to {clean_rename_db}."
                        )
                        st.rerun()
                else:
                    # Proceed to interact with the database
                    interact_with_database(selected_db)
        else:
            st.write("No databases found.")
    else:
        st.write("No databases found.")


if __name__ == "__main__":
    main()
